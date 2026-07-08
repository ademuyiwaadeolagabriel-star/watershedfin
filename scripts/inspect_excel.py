#!/usr/bin/env python3
"""Inspect the Excel CAM file thoroughly — list all sheets and dump key cells."""
import openpyxl
import warnings
warnings.filterwarnings("ignore")

FILE = "/home/z/my-project/upload/BLESSED ONYEKACHI ELECTRONICS - Final Approval.xlsx"

wb = openpyxl.load_workbook(FILE, data_only=False)
print("=" * 80)
print(f"FILE: {FILE}")
print(f"TOTAL SHEETS: {len(wb.sheetnames)}")
print("=" * 80)
print("\nSHEET NAMES (in order):")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"  {i:2d}. {name!r}  ({ws.max_row} rows x {ws.max_column} cols)")

# Also load with data_only to get cached values
wb_v = openpyxl.load_workbook(FILE, data_only=True)

print("\n" + "=" * 80)
print("SHEET-BY-SHEET DUMP (first 60 rows, first 12 cols)")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_v = wb_v[sheet_name]
    print("\n" + "#" * 80)
    print(f"### SHEET: {sheet_name!r}  ({ws.max_row}r x {ws.max_column}c)")
    print("#" * 80)

    max_row = min(ws.max_row, 80)
    max_col = min(ws.max_column, 12)

    for r in range(1, max_row + 1):
        row_cells = []
        has_content = False
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            val = cell.value
            val_v = cell_v.value
            if val is None and val_v is None:
                row_cells.append("")
            else:
                has_content = True
                # prefer the formula if there is one, then cached value
                if isinstance(val, str) and val.startswith("="):
                    row_cells.append(f"{val}")
                elif val is not None:
                    s = str(val)
                    if len(s) > 30:
                        s = s[:27] + "..."
                    row_cells.append(s)
                else:
                    s = str(val_v) if val_v is not None else ""
                    if isinstance(val_v, float):
                        s = f"{val_v:.2f}"
                    if len(s) > 30:
                        s = s[:27] + "..."
                    row_cells.append(s)
        if has_content:
            print(f"  R{r:3d} | " + " | ".join(row_cells))
