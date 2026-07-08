#!/usr/bin/env python3
"""Deep dive into the FINANCIAL ANALYSIS sheet — dump all 239 rows by 15 cols."""
import openpyxl
import warnings
warnings.filterwarnings("ignore")

FILE = "/home/z/my-project/upload/BLESSED ONYEKACHI ELECTRONICS - Final Approval.xlsx"

wb = openpyxl.load_workbook(FILE, data_only=False)
wb_v = openpyxl.load_workbook(FILE, data_only=True)

# Full FINANCIAL ANALYSIS sheet
print("=" * 100)
print("FULL DUMP: FINANCIAL ANALYSIS SHEET (239 rows x 15 cols)")
print("=" * 100)
ws = wb["FINANCIAL ANALYSIS"]
ws_v = wb_v["FINANCIAL ANALYSIS"]
for r in range(1, ws.max_row + 1):
    row_cells = []
    has_content = False
    for c in range(1, 16):
        cell = ws.cell(row=r, column=c)
        cell_v = ws_v.cell(row=r, column=c)
        val = cell.value
        val_v = cell_v.value
        if val is None and val_v is None:
            row_cells.append("")
        else:
            has_content = True
            if isinstance(val, str) and val.startswith("="):
                row_cells.append(f"F:{val[:40]}")
            elif val is not None:
                s = str(val)
                if len(s) > 35:
                    s = s[:32] + "..."
                row_cells.append(s)
            else:
                if isinstance(val_v, float):
                    s = f"V:{val_v:.2f}"
                else:
                    s = f"V:{val_v}"
                if len(s) > 35:
                    s = s[:32] + "..."
                row_cells.append(s)
    if has_content:
        print(f"R{r:3d} | " + " | ".join(row_cells))

# Also dump COMMITTEE'S DECISION sheet and LO/BM visitation
for sheet_name in ["LO VISITATION REPORT", "BM VISITATION REPORT", "COLLATERAL PLEDGE", "COMMITTEE'S DECISION"]:
    print("\n" + "=" * 100)
    print(f"SHEET: {sheet_name}")
    print("=" * 100)
    ws = wb[sheet_name]
    ws_v = wb_v[sheet_name]
    max_c = min(ws.max_column, 17)
    for r in range(1, ws.max_row + 1):
        row_cells = []
        has_content = False
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            val = cell.value
            val_v = cell_v.value
            if val is None and val_v is None:
                row_cells.append("")
            else:
                has_content = True
                if isinstance(val, str) and val.startswith("="):
                    row_cells.append(f"F:{val[:30]}")
                elif val is not None:
                    s = str(val)
                    if len(s) > 28:
                        s = s[:25] + "..."
                    row_cells.append(s)
                else:
                    if isinstance(val_v, float):
                        s = f"{val_v:.0f}"
                    else:
                        s = str(val_v)[:28]
                    row_cells.append(s)
        if has_content:
            print(f"R{r:3d} | " + " | ".join(row_cells))
